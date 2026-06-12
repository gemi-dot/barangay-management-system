import openpyxl
from django.core.management.base import BaseCommand, CommandError
from residents.models import Resident


DEFAULT_DATE_OF_BIRTH = '1900-01-01'
DEFAULT_GENDER = 'M'


class Command(BaseCommand):
    help = 'Import residents from an Excel (.xlsx) file with update-safe matching'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Path to the Excel file to import')
        parser.add_argument(
            '--sheet',
            type=str,
            default='Residents Import',
            help='Sheet name to import from (default: Residents Import)'
        )
        parser.add_argument(
            '--skip-errors',
            action='store_true',
            help='Skip rows with errors instead of stopping'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        sheet_name = options['sheet']
        skip_errors = options['skip_errors']
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(excel_file)
            
            if sheet_name not in wb.sheetnames:
                raise CommandError(f'Sheet "{sheet_name}" not found. Available sheets: {", ".join(wb.sheetnames)}')
            
            ws = wb[sheet_name]
            
            # Get headers from row 1
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            if not headers or all(h is None for h in headers):
                raise CommandError('No headers found in row 1')
            
            created_count = 0
            updated_count = 0
            error_count = 0
            skipped_count = 0
            
            # Process data starting from row 4 (skip header rows 1-3)
            for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=False), start=4):
                # Check if row is empty
                if all(cell.value is None for cell in row):
                    continue
                
                try:
                    # Create dictionary from headers and values
                    row_dict = {}
                    for col_num, header in enumerate(headers):
                        if header:
                            value = row[col_num].value if col_num < len(row) else None
                            row_dict[header] = value if value is not None else ''
                    
                    # Prepare data with proper type conversions
                    data = {
                        'first_name': str(row_dict.get('first_name', '')).strip(),
                        'middle_name': str(row_dict.get('middle_name', '')).strip() or '',
                        'last_name': str(row_dict.get('last_name', '')).strip(),
                        'suffix': str(row_dict.get('suffix', '')).strip() or '',
                        'gender': str(row_dict.get('gender', 'M')).strip(),
                        'date_of_birth': row_dict.get('date_of_birth', ''),
                        'place_of_birth': str(row_dict.get('place_of_birth', '')).strip(),
                        'contact_number': str(row_dict.get('contact_number', '')).strip() or '',
                        'email': str(row_dict.get('email', '')).strip() or '',
                        'civil_status': str(row_dict.get('civil_status', 'single')).strip(),
                        'citizenship': str(row_dict.get('citizenship', 'FILIPINO')).strip() or 'FILIPINO',
                        'house_number': str(row_dict.get('house_number', '')).strip() or None,
                        'street': str(row_dict.get('street', '')).strip() or None,
                        'zone': str(row_dict.get('zone', 'Purok Talisay')).strip(),
                        'barangay': str(row_dict.get('barangay', 'ABGAO')).strip() or 'ABGAO',
                        'city_municipality': str(row_dict.get('city_municipality', 'MAASIN')).strip() or 'MAASIN',
                        'province': str(row_dict.get('province', 'SOUTHERN LEYTE')).strip() or 'SOUTHERN LEYTE',
                        'zip_code': str(row_dict.get('zip_code', '6600')).strip() or '6600',
                        'educational_attainment': str(row_dict.get('educational_attainment', '')).strip() or None,
                        'employment_status': str(row_dict.get('employment_status', '')).strip() or None,
                        'occupation': str(row_dict.get('occupation', '')).strip() or '',
                        'father_name': str(row_dict.get('father_name', '')).strip() or '',
                        'mother_name': str(row_dict.get('mother_name', '')).strip() or '',
                        'spouse_name': str(row_dict.get('spouse_name', '')).strip() or '',
                        'emergency_contact_name': str(row_dict.get('emergency_contact_name', '')).strip() or None,
                        'emergency_contact_number': str(row_dict.get('emergency_contact_number', '')).strip() or None,
                        'emergency_contact_relationship': str(row_dict.get('emergency_contact_relationship', '')).strip() or None,
                        'voters_id': str(row_dict.get('voters_id', '')).strip() or '',
                        'philhealth_number': str(row_dict.get('philhealth_number', '')).strip() or '',
                        'sss_gsis_number': str(row_dict.get('sss_gsis_number', '')).strip() or '',
                        'tin_number': str(row_dict.get('tin_number', '')).strip() or '',
                        'precinct_number': str(row_dict.get('precinct_number', '')).strip() or '',
                        'pwd_type': str(row_dict.get('pwd_type', '')).strip() or '',
                        'blood_type': str(row_dict.get('blood_type', '')).strip() or '',
                        'allergies': str(row_dict.get('allergies', '')).strip() or '',
                        'medical_conditions': str(row_dict.get('medical_conditions', '')).strip() or '',
                    }
                    
                    # Handle monthly_income conversion
                    try:
                        monthly_income = row_dict.get('monthly_income', '')
                        if monthly_income:
                            data['monthly_income'] = float(monthly_income)
                        else:
                            data['monthly_income'] = None
                    except (ValueError, TypeError):
                        data['monthly_income'] = None
                    
                    # Handle boolean fields
                    for bool_field in ['is_pwd', 'is_senior_citizen', 'is_solo_parent', 'is_indigenous', 'is_4ps_beneficiary']:
                        bool_val = str(row_dict.get(bool_field, 'false')).strip().lower()
                        data[bool_field] = bool_val in ['true', '1', 'yes']
                    
                    # Convert date objects to strings if needed
                    if data['date_of_birth']:
                        import datetime
                        if isinstance(data['date_of_birth'], datetime.date):
                            data['date_of_birth'] = data['date_of_birth'].strftime('%Y-%m-%d')
                        else:
                            data['date_of_birth'] = str(data['date_of_birth']).strip()

                    if not data['date_of_birth']:
                        data['date_of_birth'] = DEFAULT_DATE_OF_BIRTH

                    if data['gender'] not in {'M', 'F'}:
                        data['gender'] = DEFAULT_GENDER
                    
                    # Validate required fields
                    if not data['first_name']:
                        raise ValueError('first_name is required')
                    if not data['last_name']:
                        raise ValueError('last_name is required')
                    
                    # Update-safe matching priority:
                    # 1) voters_id when present
                    # 2) first_name + last_name + date_of_birth fallback
                    existing = None
                    voters_id = data.get('voters_id', '').strip()
                    if voters_id:
                        existing = Resident.objects.filter(voters_id__iexact=voters_id).first()

                    if existing is None:
                        existing = Resident.objects.filter(
                            first_name__iexact=data['first_name'],
                            last_name__iexact=data['last_name'],
                            date_of_birth=data['date_of_birth'],
                        ).first()

                    if existing is not None:
                        for field, value in data.items():
                            setattr(existing, field, value)
                        existing.save()
                        updated_count += 1
                    else:
                        Resident.objects.create(**data)
                        created_count += 1
                    
                except Exception as e:
                    error_count += 1
                    error_msg = f"Row {row_num}: {str(e)}"
                    
                    if skip_errors:
                        self.stdout.write(self.style.WARNING(f'SKIPPED - {error_msg}'))
                        skipped_count += 1
                    else:
                        raise CommandError(error_msg)
            
            # Print summary
            self.stdout.write(self.style.SUCCESS(
                f'\n✓ Import complete: created {created_count}, updated {updated_count}'
            ))
            if skipped_count > 0:
                self.stdout.write(self.style.WARNING(f'⚠ Skipped {skipped_count} rows due to errors'))
            if error_count > 0 and not skip_errors:
                self.stdout.write(self.style.ERROR(f'✗ {error_count} error(s) encountered'))
                    
        except FileNotFoundError:
            raise CommandError(f'Excel file not found: {excel_file}')
        except Exception as e:
            raise CommandError(f'Error reading Excel file: {str(e)}')
