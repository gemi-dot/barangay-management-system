import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.core.management.base import BaseCommand


class Command(BaseCommand):
    help = 'Generate an Excel template for resident data import'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output',
            type=str,
            default='data/residents_import_template.xlsx',
            help='Output file path for the Excel template'
        )

    def handle(self, *args, **options):
        output_file = options['output']
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Residents Import"
        
        # Define field headers with their properties
        fields = [
            # Personal Information
            ('first_name', 'Text', '*Required'),
            ('middle_name', 'Text', 'Optional'),
            ('last_name', 'Text', '*Required'),
            ('suffix', 'Text', 'Optional'),
            ('gender', 'Choice: M/F', '*Required'),
            ('date_of_birth', 'Date (YYYY-MM-DD)', '*Required'),
            ('place_of_birth', 'Text', 'Optional'),
            
            # Contact Information
            ('contact_number', 'Phone', 'Optional'),
            ('email', 'Email', 'Optional'),
            
            # Civil Status & Citizenship
            ('civil_status', 'Choice: single/married/widowed/separated/divorced', 'Optional'),
            ('citizenship', 'Text', 'Default: FILIPINO'),
            
            # Address Information
            ('house_number', 'Text', 'Optional'),
            ('street', 'Text', 'Optional'),
            ('zone', 'Choice: Purok*', 'Optional'),
            ('barangay', 'Text', 'Default: ABGAO'),
            ('city_municipality', 'Text', 'Default: MAASIN'),
            ('province', 'Text', 'Default: SOUTHERN LEYTE'),
            ('zip_code', 'Text', 'Default: 6600'),
            
            # Educational & Employment
            ('educational_attainment', 'Choice: See Sheet2', 'Optional'),
            ('employment_status', 'Choice: See Sheet2', 'Optional'),
            ('occupation', 'Text', 'Optional'),
            ('monthly_income', 'Decimal', 'Optional'),
            
            # Family Information
            ('father_name', 'Text', 'Optional'),
            ('mother_name', 'Text', 'Optional'),
            ('spouse_name', 'Text', 'Optional'),
            ('emergency_contact_name', 'Text', 'Optional'),
            ('emergency_contact_number', 'Phone', 'Optional'),
            ('emergency_contact_relationship', 'Text', 'Optional'),
            
            # Government IDs
            ('voters_id', 'Text', 'Optional'),
            ('philhealth_number', 'Text', 'Optional'),
            ('sss_gsis_number', 'Text', 'Optional'),
            ('tin_number', 'Text', 'Optional'),
            ('precinct_number', 'Text', 'Optional'),
            
            # Health & Special Categories
            ('is_pwd', 'Boolean: true/false', 'Optional'),
            ('pwd_type', 'Text', 'Optional'),
            ('is_senior_citizen', 'Boolean: true/false', 'Optional'),
            ('is_solo_parent', 'Boolean: true/false', 'Optional'),
            ('is_indigenous', 'Boolean: true/false', 'Optional'),
            ('is_4ps_beneficiary', 'Boolean: true/false', 'Optional'),
            
            # Health Information
            ('blood_type', 'Text', 'Optional'),
            ('allergies', 'Text', 'Optional'),
            ('medical_conditions', 'Text', 'Optional'),
        ]
        
        # Styling
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        required_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        optional_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        default_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, (field_name, field_type, field_status) in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Write field information rows
        ws.insert_rows(2)
        ws.insert_rows(3)
        
        for col_num, (field_name, field_type, field_status) in enumerate(fields, 1):
            # Type row
            cell = ws.cell(row=2, column=col_num)
            cell.value = field_type
            cell.font = Font(size=9, italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            
            # Status row
            cell = ws.cell(row=3, column=col_num)
            cell.value = field_status
            cell.font = Font(size=9, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            
            if '*Required' in field_status:
                cell.fill = required_fill
            elif 'Default' in field_status:
                cell.fill = default_fill
            else:
                cell.fill = optional_fill
        
        # Add sample data rows (starting at row 4)
        sample_data = [
            [
                'Juan', '', 'Dela Cruz', 'Jr.', 'M', '1985-05-15', 'Maasin City',
                '09171234567', 'juan@email.com', 'married', 'FILIPINO',
                '123', 'Main Street', 'Purok Talisay', 'ABGAO', 'MAASIN', 'SOUTHERN LEYTE', '6600',
                'college', 'employed', 'Manager', '25000',
                'Jose Dela Cruz', 'Maria Santos', 'Ana Rodriguez',
                'Maria Dela Cruz', '09281234567', 'Sister',
                'V-12345678', 'PH-98765432', 'SSS-12-1234567-1', 'TIN-123-456-789', '001',
                'false', '', 'false', 'false', 'false', 'false',
                'O+', '', ''
            ],
            [
                'Maria', 'Rose', 'Santos', '', 'F', '1990-08-22', 'Maasin City',
                '09281234567', 'maria@email.com', 'single', 'FILIPINO',
                '456', 'Rizal Street', 'Purok Malunggay', 'ABGAO', 'MAASIN', 'SOUTHERN LEYTE', '6600',
                'high_school', 'unemployed', 'Housewife', '0',
                'Pedro Santos', 'Rosa Garcia', '',
                'Juan Dela Cruz', '09171234567', 'Brother',
                'V-87654321', 'PH-11223344', 'SSS-12-9876543-2', 'TIN-987-654-321', '002',
                'false', '', 'false', 'false', 'false', 'true',
                'A-', 'Shellfish', ''
            ],
        ]
        
        for row_num, row_data in enumerate(sample_data, 4):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
        
        # Add empty rows for user data (rows 6-25)
        for row_num in range(6, 26):
            for col_num in range(1, len(fields) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Set column widths
        for col_num in range(1, len(fields) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Freeze the header and info rows
        ws.freeze_panes = 'A4'
        
        # Create a reference sheet
        ref_ws = wb.create_sheet("Field Reference")
        
        ref_ws['A1'] = "Field Reference Guide"
        ref_ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Educational Attainment Choices
        ref_ws[f'A{row}'] = "Educational Attainment"
        ref_ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ref_ws[f'A{row}'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        ref_ws[f'B{row}'] = "Code"
        ref_ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ref_ws[f'B{row}'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        row += 1
        for code, display in [
            ('no_formal', 'No Formal Education'),
            ('elementary', 'Elementary'),
            ('high_school', 'High School'),
            ('vocational', 'Vocational'),
            ('college', 'College'),
            ('post_graduate', 'Post Graduate'),
        ]:
            ref_ws[f'A{row}'] = display
            ref_ws[f'B{row}'] = code
            row += 1
        
        # Employment Status Choices
        row += 1
        ref_ws[f'A{row}'] = "Employment Status"
        ref_ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ref_ws[f'A{row}'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        ref_ws[f'B{row}'] = "Code"
        ref_ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ref_ws[f'B{row}'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        row += 1
        for code, display in [
            ('employed', 'Employed'),
            ('unemployed', 'Unemployed'),
            ('student', 'Student'),
            ('retired', 'Retired'),
            ('self_employed', 'Self Employed'),
            ('ofw', 'OFW'),
        ]:
            ref_ws[f'A{row}'] = display
            ref_ws[f'B{row}'] = code
            row += 1
        
        # Zone Choices
        row += 1
        ref_ws[f'A{row}'] = "Zones (Puroks)"
        ref_ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
        ref_ws[f'A{row}'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        
        row += 1
        for zone in [
            'Purok Talisay',
            'Purok Malunggay',
            'Purok Mancinitas',
            'Purok Narra',
            'Purok Kulo',
            'Purok Ipil-ipil',
            'Purok Tugas',
        ]:
            ref_ws[f'A{row}'] = zone
            row += 1
        
        # Set column widths for reference sheet
        ref_ws.column_dimensions['A'].width = 30
        ref_ws.column_dimensions['B'].width = 25
        
        # Save the workbook
        wb.save(output_file)
        
        self.stdout.write(self.style.SUCCESS(f'✓ Excel template created: {output_file}'))
        self.stdout.write(self.style.SUCCESS(f'  - Sheet 1: "Residents Import" with field headers and sample data'))
        self.stdout.write(self.style.SUCCESS(f'  - Sheet 2: "Field Reference" with choice field values'))
